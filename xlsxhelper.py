import openpyxl
import io
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from database import get_sprav_result, get_vedomosti, get_students, update_result

def generate_xlsx_ved(id):
    vedomost = get_vedomosti(id)
    header_data = [
        ["ID", vedomost[0]["id"]],
        ["Facultet name", vedomost[0]["f_name"]],
        ["Group ID", vedomost[0]["group_id"]],
        ["Group name", vedomost[0]["g_name"]],
        ["Discipline", vedomost[0]["discipline"]],
        ["Type", vedomost[0]["control_type"]],
        ["Teacher", vedomost[0]["name_of_ekzaminator"]],
        ["Hours", vedomost[0]["hours_total"]],
        ["Date", vedomost[0]["control_date"]],
    ]
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    row_number = 1
    for header_row in header_data:
        worksheet[f"A{row_number}"] = header_row[0]
        worksheet[f"C{row_number}"] = header_row[1]
        worksheet.merge_cells(f"A{row_number}:B{row_number}")
        worksheet.merge_cells(f"C{row_number}:F{row_number}")
        worksheet[f"A{row_number}"].alignment = Alignment(horizontal="right")
        worksheet[f"C{row_number}"].alignment = Alignment(horizontal="center")
        row_number += 1
    # Create a list of choices
    choices = [""]
    results = get_sprav_result()
    for res in results:
        choices.append(res["name"])

    # Create a Data Validation rule
    dv = DataValidation(
        type="list", formula1=f'"{",".join(choices)}"', allow_blank=True
    )

    # dv = DataValidation(type="list", formula1='"Zachteno,Horosho,Otlichno"', allow_blank=True)

    # Add the Data Validation rule to the worksheet
    worksheet.add_data_validation(dv)

    students = get_students(id)
    print(id)
    print(students)
    row_number += 1
    worksheet[f"A{row_number}"] = "Код ведомости"
    worksheet[f"B{row_number}"] = "Студномер"
    worksheet[f"C{row_number}"] = "ФИО"
    worksheet[f"D{row_number}"] = "Результат"
    row_number += 1
    stud_number = 1
    for student in students:
        worksheet[f"A{row_number}"] = stud_number
        worksheet[f"B{row_number}"] = student["student_number"]
        worksheet[f"C{row_number}"] = student["full_name"]
        worksheet[f"D{row_number}"] = ""
        dv.add(worksheet[f"D{row_number}"])
        row_number += 1
        stud_number += 1

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def upload_xlsx_ved(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.active
    data = worksheet.values
    # process data
    for row in data:
        return update_result(ved_id=worksheet["B1"], stud_id=row[1], result=row[3])
